from io import StringIO
import random
import numpy
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from scipy.stats import wasserstein_distance
from openpyxl import Workbook
import os

numpy.set_printoptions(legacy='1.25')

FILE_NAME = "" # global variable to be filled by runFromFile()
# Find median size of blocks
def findMedian(total:int, blockCount):
    GENOME_SIZE = int(jumpgenomesize.get())

    if total == 0:
        return 0

    median = total / 2.0
    # If the total is odd
    if total % 2 == 1:
        for i in range(GENOME_SIZE):
            median -= blockCount[i]
            if median <= 0:
                return i + 1

    # Else the total is even
    for i in range(GENOME_SIZE):
        median -= blockCount[i]
        if median < 0.500000000001:
            i += 1
            if median > -0.499999999999:
                for j in range(GENOME_SIZE):
                    if blockCount[j] != 0:
                        return (j + 1 + i) / 2.0
            return i

    return -1

# Display results in GUI
def display_results(blockCount):
    genomeSize = len(blockCount)
    # Calculate the total amount of blocks
    totalBlocks = 0
    for i in range(genomeSize):
        totalBlocks += blockCount[i]
    #print(totalBlocks)

    # Find block frequency
    frequency = [0] * genomeSize
    if totalBlocks > 0:
        for i in range(genomeSize):
            frequency[i] = blockCount[i] / totalBlocks

    # Find average
    avg = 0.0
    if totalBlocks > 0:
        for i in range(genomeSize):
            avg += blockCount[i]*(i+1)
        avg /= totalBlocks

    for widget in frame_right.winfo_children():
        widget.destroy()

    # Labels for average and median
    label_avg = tk.Label(frame_right, text=f"Average Block Length: {avg:.2f}", font=("Arial", 12))
    label_avg.pack(pady=5)
    label_median = tk.Label(frame_right, text=f"Median Block Length: {findMedian(totalBlocks, blockCount)}", font=("Arial", 12))
    label_median.pack(pady=5)

    # Formula from preposition 4
    P = numpy.exp((-1) * float(edgelength.get()))
    Q = 1-P
    formulaDistribution = [0] * genomeSize
    Z = Q + 4*(P**2)*Q/(1+P)
    formulaDistribution[0] = (Q + 4*(P**2)*(Q**2))/Z
    for k in range(2, genomeSize):
        formulaDistribution[k-1] = (4 * (P**(2*k)) * Q**2)/Z
    formulaDistribution[genomeSize-1] = 1 - sum(formulaDistribution)
    # print("Formula distribution:")
    # print(formulaDistribution)

    label_wasserstein = tk.Label(frame_right, text=f"Wasserstein Distance: {compareDistributions(frequency, formulaDistribution)}", font=("Arial", 12))
    label_wasserstein.pack(pady=5)

    # Prepare data for histogram
    lengths = [(i+1) for i in range(genomeSize) if blockCount[i] > 0]
    valuesReal = [frequency[i] for i in range(genomeSize) if frequency[i] > 0]
    valuesSim = [formulaDistribution[i] for i in range(genomeSize) if frequency[i] > 0]

    # Plot with matplotlib
    fig, ax = plt.subplots(figsize=(6, 4))
    ax.plot(lengths, valuesSim, color='red', label='Simulated data', marker='s')   # Red line with squares
    ax.plot(lengths, valuesReal, color='blue', label='Induced data', marker='o')   # Blue line with circles
    ax.set_title("Synteny block distribution")
    ax.set_xlabel("Synteny Block Length")
    ax.set_ylabel("Frequency")

    # Embed plot into Tkinter
    canvas = FigureCanvasTkAgg(fig, master=frame_right)
    canvas.draw()
    canvas.get_tk_widget().pack()

# Print all blocks
def printBlocks(blocks):
    for block in blocks:
        print(str(block[0]) + ": " + str(block[1]) + "-" + str(block[1]+block[3]-1) + ", " + str(block[2]) + "-" + str(block[2]+block[3]-1))

# Jump model
def jump(genome, edgelength):
    GENOME_SIZE = int(jumpgenomesize.get())
    # Determine how many jumps should occur based on a poisson distribution
    jumpcount = numpy.random.poisson(edgelength*GENOME_SIZE)

    # Perform the jumps
    output = genome.copy()
    for i in range(jumpcount):
        # Determine who jumps and to where
        who = random.randint(0,GENOME_SIZE-1)
        where = random.randint(0,GENOME_SIZE-1)
        # Reroll if it jumped to the exact same spot
        while (output[where] == who):
            where = random.randint(0,GENOME_SIZE-1)
        output.insert(where,output.pop(who))

    #print("Total jumps: " + str(jumpcount) + " with edge length " + str(edgelength))
    return output

# Find synteny blocks between 2 genomes in the jump model
def findSyntenyJump(genome1, genome2):
    GENOME_SIZE = int(jumpgenomesize.get())
    blockCount = [0] * GENOME_SIZE
    # block = []
    # blocks = []

    # Iterate through the genomes
    for i in range(GENOME_SIZE):
        length = 0
        longestBlockLength = 0
        # block = []

        # Skip deleted genes
        while (i < GENOME_SIZE and genome1[i] == -1):
            i += 1
        if (i == GENOME_SIZE):
            break
        
        for j in range(GENOME_SIZE):
            # First gene in block found, find block length
            if i < GENOME_SIZE and genome1[i] != -1 and genome1[i] == genome2[j]:
                length = 1
                while (i+length < GENOME_SIZE and j+length < GENOME_SIZE and genome1[i+length] != -1 and genome1[i+length] == genome2[j+length]):
                    length += 1

                # If the current block is longer than the previously longest block, replace it
                if (length > longestBlockLength):
                    longestBlockLength = length
                    startPos = j
                j = j+length

        # Add the block to the list of blocks
        if (longestBlockLength > 0):
            for k in range(longestBlockLength):
                # block.append(genome1[i+k])
                genome1[i+k] = -1
                genome2[startPos+k] = -1
            # blocks.append([block, i, startPos, length]) 
            blockCount[length-1] += 1

    return blockCount

# Find synteny blocks between 2 real genomes from input
def findSyntenyReal(genome1, genome2):
    genome1size = len(genome1)
    genome2size = len(genome2)
    blockCount = [0] * min(genome1size,genome2size)
    # block = []
    blocks = []

    # Because the genomes are circular, move the last gene in the first genome to position 0 until it is no longer the start of a block
    push = 0
    j = 0
    while j < genome2size:
        if push < min(genome1size,genome2size) and j > -1 and genome2[j] == genome1[-1] and genome2[(j+1)%genome2size] == genome1[0]:
            if push == 0:
                startPos2 = (j+1)%genome2size
            push += 1
            genome1.insert(0, genome1.pop())
            j -= 2
        j += 1

    # If the genomes are the same, return early
    if min(genome1size,genome2size) == push:
        blocks.append([push,0,genome1size-1,startPos2,(startPos2-1)%genome2size])
        blockCount[min(genome1size,genome2size)-1] = 1
        return blocks

    # Iterate through the genomes
    longestBlockLength = 1
    while longestBlockLength > 0:
        longestBlockLength = 0
        for i in range(genome1size):
            length = 0

            # Skip deleted genes
            while (i < genome1size and genome1[i] == -1):
                i += 1
            if (i == genome1size): # reached end of genome
                break
            
            # Compare with second genome (regular form)
            for j in range(genome2size):
                # First gene in block found, find block length
                if i < genome1size and genome1[i] != -1 and genome1[i] == genome2[j]:
                    length = 1
                    while (genome1[(i+length)%genome1size] != -1 and genome1[(i+length)%genome1size] == genome2[(j+length)%genome2size]):
                        length += 1

                    # If the current block is longer than the previous longest block, replace it
                    if (length > longestBlockLength):
                        inversed = False
                        longestBlockLength = length
                        startPos1 = i
                        startPos2 = j
                    j = j+length
            
            # Compare with second genome (inversed form)
            for j in range(genome2size-1, 0, -1):
                # First gene in block found, find block length
                if i < genome1size and genome1[i] != -1 and genome1[i] == genome2[j]:
                    length = 1
                    while (length < genome1size and genome1[(i+length)%genome1size] != -1 and genome1[(i+length)%genome1size] == genome2[(j-length)%genome2size]):
                        length += 1

                    # If the current block is longer than the previous longest block, replace it
                    if (length > longestBlockLength):
                        inversed = True
                        longestBlockLength = length
                        startPos1 = i
                        startPos2 = j
                    j = j-length

        # Add the longest block found to the list of blocks
        if (longestBlockLength > 0):
            blockCount[longestBlockLength-1] += 1
            for k in range(longestBlockLength):
                #block.append(genome1[startPos1+k])
                genome1[(startPos1+k)%genome1size] = -1
                if (inversed): 
                    genome2[(startPos2-k)%genome2size] = -1
                else:
                    genome2[(startPos2+k)%genome2size] = -1

            if (inversed):
                blocks.append([-longestBlockLength, (startPos1-push)%genome1size, (startPos1+longestBlockLength-push-1)%genome1size,
                               startPos2, (startPos2-longestBlockLength+1)%genome2size])
            else:
                blocks.append([longestBlockLength, (startPos1-push)%genome1size, (startPos1+longestBlockLength-push-1)%genome1size,
                               startPos2, (startPos2+longestBlockLength-1)%genome2size]) 
            #blockCount[longestBlockLength-1] += 1

    return blocks, blockCount

# Convert block count to frequency count
def normalizeMatrix(blockCount):
    # Delete useless entries
    while (blockCount[-1] == 0):
        blockCount.pop()

    # Calculate the total amount of blocks
    totalBlocks = 0
    for i in range(len(blockCount)):
        totalBlocks += blockCount[i]
    #print(totalBlocks)

    if totalBlocks > 0:
        for i in range(len(blockCount)):
            blockCount[i] /= totalBlocks

# Compare 2 distributions
def compareDistributions(distributionReal, distributionSimulation):
    # Sort the real distribution into bins as equally as possible (max 10 bins)
    bins = [0]
    bestBins = bins
    binMaxLength = [0] # The maximum block length (minus 1 because arrays start at 0) found in the bin i
    bestBinMaxLength = binMaxLength
    # Try 10 bins, then 9 bins, then 8... until 2 bins and find the number that gives the best ratio, give some leeway to bigger bin counts
    for binCount in range(10, 1, -1):
        # Reset the current array of bins
        bins = [0] * binCount
        binMaxLength = [-1] * binCount
        # Handle bins
        for i in range(0, binCount):
            for j in range(max(binMaxLength)+1, len(distributionReal)):
                bins[i] += distributionReal[j]
                binMaxLength[i] = j
                if(bins[i] >= 1/binCount):
                    break
            
        if (min(bestBins) == 0 or (min(bins) > 0 and max(bins)/min(bins) < (max(bestBins)/min(bestBins)) - 0.01)):
            bestBins = bins
            bestBinMaxLength = binMaxLength
    
    # Remove bins of capacity 0
    for i in range(len(bestBins)-1, 0, -1):
        if (bestBins[i] == 0):
            bestBins.pop(i)
            bestBinMaxLength.pop(i)

    # Sort the simulated distribution into bins following the ranges we got from the real data
    binsSimulation = [0] * len(bestBins)
    for j in range(0, bestBinMaxLength[0]+1):
        binsSimulation[0] += distributionSimulation[j]
    for i in range(1, len(bestBins)):
        for j in range(bestBinMaxLength[i-1]+1, bestBinMaxLength[i]+1):
            binsSimulation[i] += distributionSimulation[j]
        # Add all remaining blocks to the last bin
        if (i == len(bestBins)-1):
            for j in range(bestBinMaxLength[i]+1, len(distributionSimulation)):
                binsSimulation[i] += distributionSimulation[j]

    # print("Synteny block lengths bin distribution:")
    # print("1-" + str(bestBinMaxLength[0]+1), end= " ")
    # for i in range(1, len(bestBins)):
    #     print(str(bestBinMaxLength[i-1]+2) + "-" + str(bestBinMaxLength[i]+1), end= " ")
    # print("Real data bins:")
    # print(bestBins)
    # print("Simulation bins:")
    # print(binsSimulation)

    return wasserstein_distance(bestBins, binsSimulation) 

# Run the simulation many times and take the average
def runSimulation():
    GENOME_SIZE = int(jumpgenomesize.get())
    TIMES = int(times.get())
    totalBlockCount = [0] * GENOME_SIZE
    for i in range(TIMES):
        # Initiate the first genome
        genome1 = [0] * GENOME_SIZE
        for j in range(GENOME_SIZE):
            genome1[j] = j
        
        # Get the second genome using the Poisson jump model
        genome2 = jump(genome1, float(edgelength.get()))

        # Compare the 2 genomes
        blockCount = findSyntenyJump(genome1,genome2)
        for j in range(GENOME_SIZE):
            totalBlockCount[j] += blockCount[j]

    display_results(totalBlockCount)

# Compare 2 manually input genomes
def compareGenomes():
    genome1 = list(map(int, str(gen1.get()).split()))
    genome2 = list(map(int, str(gen2.get()).split()))

    # Clean singletons
    genomeCompare1 = [0] * (len(genome1) + len(genome2))
    genomeCompare2 = [0] * (len(genome1) + len(genome2))
    for i in range(max(len(genome1), len(genome2))):
        if (i < len(genome1)):
            genomeCompare1[genome1[i]-1] = 1
        if (i < len(genome2)):
            genomeCompare2[genome2[i]-1] = 1

    for i in range(len(genomeCompare1)):
        # Gene exists in genome1 but not in genome2
        if (genomeCompare1[i] > genomeCompare2[i]):
            for j in range(len(genome1)-1, -1, -1):
                if (genome1[j] == i+1):
                    genome1.pop(j)
        # Gene exists in genome2 but not in genome1
        elif (genomeCompare1[i] < genomeCompare2[i]):
            for j in range(len(genome2)-1, -1, -1):
                if (genome2[j] == i+1):
                    genome2.pop(j)

    print("Comparing genomes:\n" + str(genome1) + "\n" + str(genome2))
    blockCount = findSyntenyReal(genome1,genome2)
    print(blockCount)

    # while (blockCount[-1] == 0):
    #     blockCount.pop()

    #display_results(blockCountRegular)

# Extract genome data
def extractGeneFromRD(line):
    splitLine = line.split()
    genomeId= splitLine[3]
    genomeCog = int(splitLine[-1][4:])
    return genomeId, genomeCog
 
# Read the file and extract genome data
def runFromFile():
    filepath = file_entry.get()
    if not filepath:
        print("No file selected.")
        return
    
    genomes = []
    print(filepath)
    global FILE_NAME
    # FILE_NAME = filepath.split("/")[:8]
    FILE_NAME = os.path.splitext(os.path.basename(filepath))[0]
    
    with open(filepath, "r") as file:
        lines = file.readlines()
        genomeNames = []
        curr_id = ""
        for line in lines:
            if line[0] == "#":
                continue
            new_id, new_num = extractGeneFromRD(line)
            if new_id != curr_id:
                genomes.append([])
                genomeNames.append(new_id)
                curr_id = new_id
            genomes[-1].append(new_num)

        # Print the matrix
        # for genome in genomes:
        #     print(genome)
        return genomes, genomeNames

# Generate graph for the ATGC from read file
def runATGC():
    genomes, genomeNames = runFromFile()
    if genomes is None:
        print("Error: Empty matrix")
        return
    
    # Make output folder
    os.makedirs("output", exist_ok=True)

    # # Initiate Excel matrix and dataframe
    # wb = Workbook()
    # ws = wb.active
    # ws.title = "Synteny Blocks"
    # for col_idx, name in enumerate(genomeNames, start=2):
    #     ws.cell(row=1, column=col_idx, value=name)
    
    # # Write row and column headers and data
    # for idx, row_name in enumerate(genomeNames, start=2):
    #     ws.cell(row=idx, column=1, value=row_name)
    #     ws.cell(row=1, column=idx, value=row_name)

    # Find max length
    genomeSize = 0
    for i in range(len(genomes)):
        if genomeSize < len(genomes[i]):
            genomeSize = len(genomes[i])

    print("starting comparison")
    # Compare every pair of genomes and write the results in Excel files
    for i in range(0, len(genomes)):
        print(i+1,"out of",len(genomes))
        for j in range(i+1, len(genomes)):
            # Compare genome i with genome j
            print("\t", j-i, "out of", len(genomes)-i-1)
            blocks, blockCount = findSyntenyReal(genomes[i].copy(),genomes[j].copy())

            # Synteny Blocks
            wb = Workbook()
            ws = wb.active
            titles = ["Length",  genomeNames[i] + " Start", genomeNames[i] + " End", genomeNames[j] + " Start", genomeNames[j] + " End"]
            ws.append(titles)
            for block in blocks:
                ws.append(block)
            wb.save(os.path.join("output", (FILE_NAME + "-" + genomeNames[i] + "-" + genomeNames[j] + "-SyntenyBlocks.xlsx")))

            # SBLD
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=1, value="Length")
            ws.cell(row=2, column=1, value="Frequency")
            normalizeMatrix(blockCount)
            for row in range(len(blockCount)):  # Start at row 2 (after header)
                ws.cell(row=row+2, column=1, value=row+1)
                ws.cell(row=row+2, column=2, value=blockCount[row])
            wb.save(os.path.join("output", (FILE_NAME + "-" + genomeNames[i] + "-" + genomeNames[j] + "-SBLD.xlsx")))

            #ws.cell(row=i+2, column=j+2, value=str(findSyntenyReal(genomes[i].copy(),genomes[j].copy())))
            #print_results(df, genomes[i], genomes[j], findSyntenyReal(genomes[i].copy(),genomes[j].copy()))
    
    print("finished")
    #wb.save('SBLD.xlsx')
    #wb.save(outputFolder / (FILE_NAME + '-sbld.xlsx'))

# Browse file
def browse_file():
    filename = filedialog.askopenfilename(title="Select Data File", filetypes=(("Pseudo-terminal utilities", "*.pty"), ("All files", "*.*")))
    if filename:
        file_entry.delete(0, tk.END)
        file_entry.insert(0, filename)

# GUI setup
root = tk.Tk()
root.title("Genome Synteny Simulation")
root.geometry("900x600")
frame_left = tk.Frame(root, padx=10, pady=10)
frame_left.pack(side=tk.LEFT, fill=tk.Y)
title_label = tk.Label(frame_left, text="Genome Synteny Simulator", font=("Arial", 16))
title_label.pack(pady=10)

# Edge length
ttk.Label(frame_left, text="Edge Length").pack()
edgelength = ttk.Entry(frame_left)
edgelength.insert(0, "0.2")  # Default value
edgelength.pack(pady=(0,30))

# Read file
ttk.Label(frame_left, text="ATGC Data File").pack()
file_entry = ttk.Entry(frame_left)
file_entry.pack()
browse_button = ttk.Button(frame_left, text="Browse", command=browse_file)
browse_button.pack()
file_sim_button = ttk.Button(frame_left, text="Find synteny blocks in file", command=runATGC)
file_sim_button.pack(pady=(10,30))

# Jump simulation
ttk.Label(frame_left, text="Number of simulations").pack()
times = ttk.Entry(frame_left)
times.insert(0, "1")  # Default value
times.pack()
ttk.Label(frame_left, text="Genome size (jump sim)").pack()
jumpgenomesize = ttk.Entry(frame_left)
jumpgenomesize.insert(0, "5000")  # Default value
jumpgenomesize.pack()
run_button = ttk.Button(frame_left, text="Run Jump Simulation", command=runSimulation)
run_button.pack(pady=(10,30))

# Direct comparison of 2 genomes
ttk.Label(frame_left, text="Genomes").pack()
gen1 = ttk.Entry(frame_left)
gen1.insert(0, "")  # Default value
gen1.pack()
gen2 = ttk.Entry(frame_left)
gen2.insert(0, "")  # Default value
gen2.pack()
run_button = ttk.Button(frame_left, text="Direct comparison", command=compareGenomes)
run_button.pack(pady=(10,0))

# Note
color_guide = tk.Label(frame_left, text="Red = formula\nBlue = induced")
color_guide.pack(pady=10)

frame_right = tk.Frame(root, padx=10, pady=10)
frame_right.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

root.mainloop()